import io
import logging
from datetime import datetime

from utils.timezone import now_local

from aiogram import Router, F
from aiogram.types import BufferedInputFile, CallbackQuery

from db import get_appointments_for_export, _price_fmt
from keyboards.inline import export_period_keyboard
from utils.admin import is_admin_callback, deny_access, IsAdminFilter
from utils.callbacks import parse_callback

logger = logging.getLogger(__name__)
router = Router()
router.message.filter(IsAdminFilter())
router.callback_query.filter(IsAdminFilter())

_PERIOD_LABEL = {
    "today": "сегодня",
    "week":  "эта_неделя",
    "month": "этот_месяц",
    "all":   "все_записи",
}

_STATUS_RU = {
    "scheduled": "Ожидает",
    "completed": "Выполнено",
    "cancelled": "Отменено",
    "no_show":   "Не пришли",
}


@router.callback_query(F.data == "admin_export")
async def cb_export_menu(callback: CallbackQuery):
    if not is_admin_callback(callback):
        await deny_access(callback)
        return
    try:
        await callback.message.edit_text(
            "📥 <b>Экспорт записей</b>\n\nВыберите период:",
            reply_markup=export_period_keyboard(),
            parse_mode="HTML",
        )
    except Exception:
        pass
    await callback.answer()


@router.callback_query(F.data.startswith("export_"))
async def cb_export_generate(callback: CallbackQuery):
    if not is_admin_callback(callback):
        await deny_access(callback)
        return

    parts = parse_callback(callback.data, "export", 1)
    if not parts:
        logger.warning("Некорректный callback: %s", callback.data)
        await callback.answer()
        return
    period = parts[0]  # today / week / month / all
    await callback.answer("Генерирую файл…")

    rows = await get_appointments_for_export(period)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Записи"

        # Заголовки
        headers = ["Дата", "Время", "Клиент", "Телефон", "Услуга", "Цена (сум)", "Статус"]
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row_idx, appt in enumerate(rows, start=2):
            fill = PatternFill("solid", fgColor="EBF1F5" if row_idx % 2 == 0 else "FFFFFF")
            values = [
                appt["date"],
                appt["time"],
                appt["name"],
                appt["phone"],
                appt["service_name"],
                appt["service_price"],
                _STATUS_RU.get(appt["status"], appt["status"]),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill

        # Ширина колонок
        widths = [12, 8, 20, 16, 25, 14, 12]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str = now_local().strftime("%Y-%m-%d")
        filename = f"записи_{_PERIOD_LABEL.get(period, period)}_{date_str}.xlsx"

        count = len(rows)
        await callback.message.answer_document(
            BufferedInputFile(buf.read(), filename=filename),
            caption=f"📥 Экспорт: {_PERIOD_LABEL.get(period, period).replace('_', ' ')}\n"
                    f"Записей: {count}",
        )

    except ImportError:
        await callback.message.answer("❌ Установи openpyxl: <code>pip install openpyxl</code>", parse_mode="HTML")
    except Exception:
        logger.exception("Ошибка генерации Excel")
        await callback.message.answer("❌ Ошибка при генерации файла.")
